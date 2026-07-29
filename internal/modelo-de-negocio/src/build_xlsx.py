"""Modelo financiero — Law Offices of Jose R. Santiago, PLLC.
Instrumento del Anexo A: se cargan los datos reales y el modelo recalcula solo."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

NAVY = "1F3864"
BAND = "F2F5F9"
YELLOW = "FFFF00"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=BAND)
IN_FILL = PatternFill("solid", fgColor=YELLOW)

F = "Arial"
BLUE = Font(name=F, size=10, color="0000FF")           # entradas
BLACK = Font(name=F, size=10)                          # fórmulas
GREEN = Font(name=F, size=10, color="008000")          # enlace a otra hoja
BOLD = Font(name=F, size=10, bold=True)
HDR = Font(name=F, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=F, size=14, bold=True, color=NAVY)
SUB = Font(name=F, size=9, italic=True, color="606060")
SECT = Font(name=F, size=11, bold=True, color=NAVY)

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%'
NUM = '#,##0;(#,##0);-'
NUM1 = '#,##0.0;(#,##0.0);-'

thin = Side(style="thin", color="C7D0DC")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def hdr_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, subtitle=None):
    ws["A1"] = text
    ws["A1"].font = TITLE
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUB


# ============================================================ INSTRUCCIONES
ws = wb.active
ws.title = "Instrucciones"
ws.sheet_view.showGridLines = False
title(ws, "Modelo financiero — Law Offices of Jose R. Santiago, PLLC",
      "Instrumento del Anexo A del documento de modelo de negocio. Versión 1.0")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 95

rows = [
    ("SECT", "Cómo usar este archivo", ""),
    ("T", "1.", "Sustituye los valores de ejemplo por los datos reales de la firma. Los valores precargados vienen del documento de modelo de negocio y son ILUSTRATIVOS: sirven para que el modelo funcione desde el primer momento, no para describir la firma."),
    ("T", "2.", "Escribe únicamente en las celdas con fondo amarillo. Todo lo demás son fórmulas y se recalcula solo."),
    ("T", "3.", "Empieza por las hojas 'Supuestos' y 'TiposDeCaso'. Con esas dos, el resto del modelo ya produce resultados."),
    ("T", "4.", "La hoja 'Palancas' muestra cuánto mueve la utilidad cada cambio operativo. Es la que hay que mirar antes de decidir dónde invertir esfuerzo."),
    ("", "", ""),
    ("SECT", "Código de color", ""),
    ("IN", "Fondo amarillo", "Celda que tú rellenas. Es la única que se edita."),
    ("BLUE", "Texto azul", "Dato de entrada escrito a mano."),
    ("BLACK", "Texto negro", "Fórmula. No tocar."),
    ("GREEN", "Texto verde", "Enlace a otra hoja del mismo archivo."),
    ("", "", ""),
    ("SECT", "Las hojas", ""),
    ("T", "Supuestos", "Costos fijos, costo por hora de cada rol, overhead por caso. Los cimientos del modelo."),
    ("T", "TiposDeCaso", "Una fila por tipo de caso: precio, horas de entrega, CAC y volumen. Aquí se ve qué producto sostiene la firma y cuál la subsidia."),
    ("T", "Embudo", "Consultas, consultas atendidas y casos firmados, mes a mes. Calcula la tasa de conversión, que es la palanca de mayor retorno."),
    ("T", "Firma", "El estado de resultados: ingreso, margen de contribución, costos fijos, utilidad y punto de equilibrio."),
    ("T", "Palancas", "Qué pasa con la utilidad si mueves conversión, horas de entrega, precio o CAC. Ordenadas por impacto."),
    ("T", "Panel", "Los cinco números del lunes por la mañana."),
    ("", "", ""),
    ("SECT", "Advertencia", ""),
    ("T", "", "Este modelo no constituye asesoría legal, fiscal ni contable. El saldo de la cuenta fiduciaria (IOLTA) NO es liquidez de la firma y no debe cargarse en ninguna celda de caja de este archivo: los fondos del cliente no devengados no son ingreso ni activo de la firma."),
]
r = 4
for kind, b, c in rows:
    if kind == "SECT":
        ws.cell(row=r, column=2, value=b).font = SECT
    elif kind == "IN":
        cell = ws.cell(row=r, column=2, value=b)
        cell.font = BOLD
        cell.fill = IN_FILL
        cell.border = BOX
        ws.cell(row=r, column=3, value=c).font = BLACK
    elif kind == "BLUE":
        ws.cell(row=r, column=2, value=b).font = BLUE
        ws.cell(row=r, column=3, value=c).font = BLACK
    elif kind == "GREEN":
        ws.cell(row=r, column=2, value=b).font = GREEN
        ws.cell(row=r, column=3, value=c).font = BLACK
    elif kind == "BLACK":
        ws.cell(row=r, column=2, value=b).font = BLACK
        ws.cell(row=r, column=3, value=c).font = BLACK
    elif kind == "T":
        ws.cell(row=r, column=2, value=b).font = BOLD
        cell = ws.cell(row=r, column=3, value=c)
        cell.font = BLACK
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
    r += 1

# ============================================================ SUPUESTOS
ws = wb.create_sheet("Supuestos")
ws.sheet_view.showGridLines = False
title(ws, "Supuestos", "Celdas amarillas = rellenar. Valores precargados ilustrativos.")
hdr_row(ws, 4, ["Concepto", "Valor", "Unidad", "Nota"], [42, 16, 14, 60])

sup = [
    ("Costo cargado por hora — abogado", 95, "$ / hora", "Salario + carga + beneficios ÷ horas trabajadas al año. NO es la tarifa que se cobra."),
    ("Costo cargado por hora — paralegal", 34, "$ / hora", "Mismo cálculo para el rol de paralegal."),
    ("Costo cargado por hora — asistente/intake", 24, "$ / hora", "Personal de recepción y gestión de expediente."),
    ("Overhead asignado por caso", 840, "$ / caso", "Costos fijos ÷ casos al año. Recalcular cada año."),
    ("Costos fijos anuales", 600000, "$ / año", "Renta, nómina no imputable a casos, software, seguros, colegiación."),
    ("Gasto de marketing mensual", 24000, "$ / mes", "Todo el gasto de captación: SEO, pago, contenido, eventos."),
    ("Meses del ejercicio", 12, "meses", "Normalmente 12. Cambiar solo para un ejercicio parcial."),
]
r = 5
for name, val, unit, note in sup:
    ws.cell(row=r, column=1, value=name).font = BOLD
    c = ws.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = IN_FILL
    c.border = BOX
    c.number_format = MONEY if "$" in unit else NUM
    ws.cell(row=r, column=3, value=unit).font = BLACK
    n = ws.cell(row=r, column=4, value=note)
    n.font = SUB
    n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

ws["A14"] = "Referencias usadas por el resto del modelo"
ws["A14"].font = SECT
ws["A15"] = "Tarifa abogado"
ws["A15"].font = BLACK
ws["B15"] = "=B5"
ws["B15"].font = BLACK
ws["B15"].number_format = MONEY2

# ============================================================ TIPOS DE CASO
ws = wb.create_sheet("TiposDeCaso")
ws.sheet_view.showGridLines = False
title(ws, "Economía por tipo de caso",
      "El corazón del modelo. Precio − CAC − costo de entrega − overhead = margen de contribución.")
heads = ["Tipo de caso", "Precio\n($)", "Horas\nabogado", "Horas\nparalegal", "CAC\n($)",
         "Casos\npor mes", "Costo de\nentrega ($)", "Overhead\n($)", "Margen por\ncaso ($)",
         "Margen\n%", "Ingreso\nanual ($)", "Margen\nanual ($)"]
hdr_row(ws, 4, heads, [30, 11, 10, 10, 10, 10, 12, 11, 13, 9, 14, 14])

casos = [
    ("Naturalización", 1800, 1.0, 3.0, 420, 12),
    ("Petición familiar (I-130)", 2600, 1.5, 4.0, 560, 10),
    ("Ajuste de estatus", 4200, 2.0, 4.5, 650, 9),
    ("Asilo afirmativo", 6500, 8.0, 10.0, 700, 4),
    ("Defensa en corte (removal)", 9000, 20.0, 12.0, 900, 3),
    ("Preparación de habeas (alcance limitado)", 2800, 1.5, 8.0, 300, 5),
    ("Cumplimiento I-9 (empleador, anual)", 7200, 6.0, 10.0, 1200, 0.5),
]
first, r = 5, 5
for name, precio, h_abg, h_par, cac, cpm in casos:
    ws.cell(row=r, column=1, value=name).font = BOLD
    for col, val, fmt in ((2, precio, MONEY), (3, h_abg, NUM1), (4, h_par, NUM1),
                          (5, cac, MONEY), (6, cpm, NUM1)):
        c = ws.cell(row=r, column=col, value=val)
        c.font = BLUE
        c.fill = IN_FILL
        c.border = BOX
        c.number_format = fmt
    # costo de entrega
    c = ws.cell(row=r, column=7, value=f"=C{r}*Supuestos!$B$5+D{r}*Supuestos!$B$6")
    c.font = BLACK; c.number_format = MONEY; c.border = BOX
    # overhead
    c = ws.cell(row=r, column=8, value="=Supuestos!$B$8")
    c.font = GREEN; c.number_format = MONEY; c.border = BOX
    # margen por caso
    c = ws.cell(row=r, column=9, value=f"=B{r}-E{r}-G{r}-H{r}")
    c.font = BOLD; c.number_format = MONEY; c.border = BOX
    # margen %
    c = ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/B{r},0)")
    c.font = BLACK; c.number_format = PCT; c.border = BOX
    # ingreso anual
    c = ws.cell(row=r, column=11, value=f"=B{r}*F{r}*Supuestos!$B$11")
    c.font = BLACK; c.number_format = MONEY; c.border = BOX
    # margen anual
    c = ws.cell(row=r, column=12, value=f"=I{r}*F{r}*Supuestos!$B$11")
    c.font = BLACK; c.number_format = MONEY; c.border = BOX
    if (r - first) % 2:
        for col in range(1, 13):
            if col not in (2, 3, 4, 5, 6):
                ws.cell(row=r, column=col).fill = BAND_FILL
    r += 1

last = r - 1
tr = r
ws.cell(row=tr, column=1, value="TOTAL").font = Font(name=F, size=10, bold=True, color=NAVY)
for col, formula, fmt in (
    (6, f"=SUM(F{first}:F{last})", NUM1),
    (9, f"=IFERROR(SUMPRODUCT(I{first}:I{last},F{first}:F{last})/SUM(F{first}:F{last}),0)", MONEY),
    (10, f"=IFERROR(L{tr}/K{tr},0)", PCT),
    (11, f"=SUM(K{first}:K{last})", MONEY),
    (12, f"=SUM(L{first}:L{last})", MONEY),
):
    c = ws.cell(row=tr, column=col, value=formula)
    c.font = Font(name=F, size=10, bold=True, color=NAVY)
    c.number_format = fmt
    c.border = Border(top=Side(style="double", color=NAVY), bottom=thin, left=thin, right=thin)
ws.cell(row=tr, column=9).comment = Comment(
    "Margen por caso ponderado por volumen, no promedio simple.", "Modelo", width=280, height=60)

ws.cell(row=tr + 2, column=1, value="Cómo leer esta tabla").font = SECT
notes = [
    "La columna 'Margen %' ordena los productos: el de arriba sostiene la firma, el de abajo puede estar subsidiándose.",
    "'Horas abogado' y 'Horas paralegal' son las horas REALES de entrega, no las facturables. Hay que medirlas aunque no se facturen: sin ellas no se sabe el margen.",
    "Bajar una hora de abogado en un tipo de caso de alto volumen vale más que subir el precio de uno de bajo volumen. La hoja 'Palancas' lo cuantifica.",
    "El overhead por caso sale de Supuestos!B8 y es igual para todos. Si un tipo de caso consume mucho más espacio o soporte, conviene asignarle un overhead propio.",
]
rr = tr + 3
for n in notes:
    c = ws.cell(row=rr, column=1, value="•  " + n)
    c.font = SUB
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
    ws.row_dimensions[rr].height = 26
    rr += 1

# ============================================================ EMBUDO
ws = wb.create_sheet("Embudo")
ws.sheet_view.showGridLines = False
title(ws, "Embudo de captación",
      "La conversión de consulta es la palanca de mayor retorno de la firma. Medirla semanalmente.")
hdr_row(ws, 4, ["Mes", "Consultas\nrecibidas", "Consultas\natendidas", "Casos\nfirmados",
                "% atendidas", "% conversión", "Gasto de\nmarketing ($)", "CAC\n($)"],
        [16, 14, 14, 13, 13, 13, 16, 12])

meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
ejemplo = [(150, 120, 42), (140, 112, 39)]
first, r = 5, 5
for i, m in enumerate(meses):
    ws.cell(row=r, column=1, value=m).font = BOLD
    vals = ejemplo[i] if i < len(ejemplo) else (None, None, None)
    for col, v in zip((2, 3, 4), vals):
        c = ws.cell(row=r, column=col, value=v)
        c.font = BLUE; c.fill = IN_FILL; c.border = BOX; c.number_format = NUM
    c = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},\"\")")
    c.font = BLACK; c.number_format = PCT; c.border = BOX
    c = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},\"\")")
    c.font = BOLD; c.number_format = PCT; c.border = BOX
    c = ws.cell(row=r, column=7, value="=Supuestos!$B$10")
    c.font = GREEN; c.number_format = MONEY; c.border = BOX
    c = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/D{r},\"\")")
    c.font = BLACK; c.number_format = MONEY; c.border = BOX
    r += 1
last = r - 1
ws.cell(row=r, column=1, value="TOTAL / PROMEDIO").font = Font(name=F, size=10, bold=True, color=NAVY)
for col, formula, fmt in (
    (2, f"=SUM(B{first}:B{last})", NUM),
    (3, f"=SUM(C{first}:C{last})", NUM),
    (4, f"=SUM(D{first}:D{last})", NUM),
    (5, f"=IFERROR(C{r}/B{r},\"\")", PCT),
    (6, f"=IFERROR(D{r}/C{r},\"\")", PCT),
    # Solo cuenta el marketing de los meses que ya tienen casos cargados, para que
    # el CAC sea correcto mientras la hoja está a medio rellenar.
    (7, f"=SUMIF(D{first}:D{last},\">0\",G{first}:G{last})", MONEY),
    (8, f"=IFERROR(G{r}/D{r},\"\")", MONEY),
):
    c = ws.cell(row=r, column=col, value=formula)
    c.font = Font(name=F, size=10, bold=True, color=NAVY)
    c.number_format = fmt
    c.border = Border(top=Side(style="double", color=NAVY), bottom=thin, left=thin, right=thin)
CONV_CELL = f"Embudo!$F${r}"

ws.cell(row=r + 2, column=1,
        value="Los tres puntos donde se pierde una consulta, por orden de frecuencia:").font = SECT
for j, t in enumerate([
    "1.  Tiempo de respuesta. Una consulta contestada en una hora convierte mucho mejor que una contestada al día siguiente.",
    "2.  Quién atiende. La persona equivocada, entre dos audiencias, con prisa.",
    "3.  Si se pide el cierre. Muchas consultas acaban en «piénselo y nos llama» en vez de en una decisión.",
]):
    c = ws.cell(row=r + 3 + j, column=1, value=t)
    c.font = SUB
    ws.merge_cells(start_row=r + 3 + j, start_column=1, end_row=r + 3 + j, end_column=8)

# ============================================================ FIRMA
ws = wb.create_sheet("Firma")
ws.sheet_view.showGridLines = False
title(ws, "Estado de resultados y punto de equilibrio",
      "Todo se alimenta de 'TiposDeCaso' y 'Supuestos'. No hay nada que rellenar aquí.")
hdr_row(ws, 4, ["Concepto", "Valor", "Nota"], [46, 18, 66])

TD_TOT = tr  # fila TOTAL de TiposDeCaso
lines = [
    ("Ingreso anual", f"=TiposDeCaso!K{TD_TOT}", MONEY,
     "Suma del ingreso de todos los tipos de caso.", False),
    ("Margen de contribución anual", f"=TiposDeCaso!L{TD_TOT}", MONEY,
     "Lo que queda tras CAC, entrega y overhead, antes de costos fijos.", False),
    ("Margen de contribución (%)", f"=IFERROR(B6/B5,0)", PCT,
     "Si baja por debajo del 40 % conviene revisar el mix de casos.", False),
    ("Costos fijos anuales", "=Supuestos!B9", MONEY, "Desde 'Supuestos'.", False),
    ("Utilidad operativa", "=B6-B8", MONEY, "Margen de contribución menos costos fijos.", True),
    ("Margen operativo (%)", "=IFERROR(B9/B5,0)", PCT, "Referencia de firma de consumidor: ~25 %.", False),
    ("", "", None, "", False),
    ("Margen por caso (ponderado)", f"=TiposDeCaso!I{TD_TOT}", MONEY,
     "Promedio ponderado por volumen, no promedio simple.", False),
    ("Casos al año para cubrir los costos fijos", "=IFERROR(B8/B12,0)", NUM,
     "Punto de equilibrio en número de casos.", True),
    ("Casos al mes para cubrir los costos fijos", "=IFERROR(B13/Supuestos!B11,0)", NUM1,
     "El número que hay que superar cada mes.", False),
    ("Casos al mes que tenemos hoy", f"=TiposDeCaso!F{TD_TOT}", NUM1, "Desde 'TiposDeCaso'.", False),
    ("Colchón sobre el equilibrio", "=IFERROR(B15/B14-1,0)", PCT,
     "Cuánto puede caer el volumen antes de entrar en pérdidas.", True),
]
r = 5
for name, formula, fmt, note, emph in lines:
    if not name:
        r += 1
        continue
    c = ws.cell(row=r, column=1, value=name)
    c.font = Font(name=F, size=10, bold=True, color=NAVY) if emph else BOLD
    v = ws.cell(row=r, column=2, value=formula)
    v.font = Font(name=F, size=11, bold=True, color=NAVY) if emph else GREEN
    v.number_format = fmt
    v.border = BOX
    if emph:
        v.fill = PatternFill("solid", fgColor="E8EEF7")
    n = ws.cell(row=r, column=3, value=note)
    n.font = SUB
    n.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

# ============================================================ PALANCAS
ws = wb.create_sheet("Palancas")
ws.sheet_view.showGridLines = False
title(ws, "Palancas — cuánto mueve cada cambio",
      "Cambia los valores amarillos y compara. Ordenadas por impacto típico, no por facilidad.")
hdr_row(ws, 4, ["Palanca", "Cambio propuesto", "Efecto por caso ($)",
                "Efecto anual ($)", "¿Cuesta dinero?"], [34, 22, 20, 20, 44])

ws["A5"] = "Escenario base (desde las otras hojas)"
ws["A5"].font = SECT
base = [
    ("Casos al año", f"=TiposDeCaso!F{TD_TOT}*Supuestos!B11", NUM1),          # B6
    ("Margen por caso hoy (ponderado)", f"=TiposDeCaso!I{TD_TOT}", MONEY),    # B7
    ("Tasa de conversión actual", f"={CONV_CELL}", PCT),                      # B8
    ("Utilidad operativa hoy", "=Firma!B9", MONEY),                           # B9
]
r = 6
for n, f_, fmt in base:
    ws.cell(row=r, column=1, value=n).font = BOLD
    c = ws.cell(row=r, column=2, value=f_)
    c.font = GREEN; c.number_format = fmt; c.border = BOX
    r += 1

r = 11
ws.cell(row=r, column=1, value="Palancas (edita la columna amarilla)").font = SECT
r += 1
hdr_row(ws, r, ["Palanca", "Cambio propuesto", "Efecto por caso ($)",
                "Efecto anual ($)", "Comentario"], [34, 22, 20, 20, 44])
pal_start = r + 1

pal = [
    ("Conversión de consulta", 0.10, PCT, None,
     f"=IFERROR($B$6*(B{pal_start}/$B$8)*$B$7,0)",
     "Sube la conversión N PUNTOS porcentuales. Actúa sobre el volumen, no sobre el margen por caso: más casos con el mismo gasto de marketing. No cuesta dinero — esa consulta ya se pagó."),
    ("Horas de abogado por caso", -1.0, NUM1,
     f"=-B{pal_start+1}*Supuestos!$B$5", f"=C{pal_start+1}*$B$6",
     "Reduce N horas de abogado por caso moviendo trabajo al paralegal o al proceso."),
    ("Horas de paralegal por caso", -1.0, NUM1,
     f"=-B{pal_start+2}*Supuestos!$B$6", f"=C{pal_start+2}*$B$6",
     "Reduce N horas de paralegal por caso mediante plantillas y automatización."),
    ("Precio", 0.08, PCT,
     f"=IFERROR(TiposDeCaso!K{TD_TOT}/$B$6*B{pal_start+3},0)", f"=C{pal_start+3}*$B$6",
     "Sube el precio un N %. Sin costo directo, pero con riesgo comercial."),
    ("CAC", -150, MONEY,
     f"=-B{pal_start+4}", f"=C{pal_start+4}*$B$6",
     "Baja el CAC N dólares desplazando la mezcla de canales hacia referidos."),
]
r = pal_start
for name, val, fmt, f_caso, f_anual, com in pal:
    ws.cell(row=r, column=1, value=name).font = BOLD
    c = ws.cell(row=r, column=2, value=val)
    c.font = BLUE; c.fill = IN_FILL; c.border = BOX; c.number_format = fmt
    if f_caso is None:
        c = ws.cell(row=r, column=3, value="n/a")
        c.font = SUB
        c.alignment = Alignment(horizontal="center")
        c.border = BOX
        c.comment = Comment(
            "Esta palanca no cambia el margen por caso: cambia cuántos casos entran. "
            "Su efecto está solo en la columna anual.", "Modelo", width=300, height=80)
    else:
        c = ws.cell(row=r, column=3, value=f_caso)
        c.font = BLACK; c.number_format = MONEY; c.border = BOX
    c = ws.cell(row=r, column=4, value=f_anual)
    c.font = BOLD; c.number_format = MONEY; c.border = BOX
    n = ws.cell(row=r, column=5, value=com)
    n.font = SUB
    n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 40
    r += 1

pal_end = r - 1
ws.cell(row=r, column=1, value="Efecto combinado sobre la utilidad").font = Font(name=F, size=10, bold=True, color=NAVY)
c = ws.cell(row=r, column=4, value=f"=SUM(D{pal_start}:D{pal_end})")
c.font = Font(name=F, size=11, bold=True, color=NAVY)
c.number_format = MONEY
c.fill = PatternFill("solid", fgColor="E8EEF7")
c.border = Border(top=Side(style="double", color=NAVY), bottom=thin, left=thin, right=thin)

r += 1
c = ws.cell(row=r, column=1, value="Utilidad operativa resultante")
c.font = Font(name=F, size=10, bold=True, color=NAVY)
c = ws.cell(row=r, column=4, value=f"=$B$9+D{r-1}")
c.font = Font(name=F, size=11, bold=True, color=NAVY)
c.number_format = MONEY
c.fill = PatternFill("solid", fgColor="E8EEF7")
c.border = BOX

r += 2
for t in [
    "El efecto de la conversión se calcula sobre el volumen: más casos con el mismo gasto de marketing y el mismo margen por caso.",
    "Las palancas se multiplican, no se suman: mejorar cinco cosas un 5 % da alrededor de +28 %, no +25 %. Este cuadro las suma de forma conservadora.",
    "La columna 'Valor propuesto' admite valores negativos. Sirve también para medir el daño: qué pasa si el CAC sube o si las horas de entrega crecen.",
]:
    c = ws.cell(row=r, column=1, value="•  " + t)
    c.font = SUB
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 24
    r += 1

# ============================================================ PANEL
ws = wb.create_sheet("Panel")
ws.sheet_view.showGridLines = False
title(ws, "Panel — los cinco números del lunes por la mañana",
      "Media hoja, una vez al mes, misma fecha. La disciplina de mirarlos vale más que la sofisticación de calcularlos.")
hdr_row(ws, 4, ["#", "Indicador", "Valor", "Referencia sana", "Qué hacer si se sale"],
        [5, 34, 18, 22, 60])

panel = [
    ("1", "Casos firmados al año", f"=TiposDeCaso!F{TD_TOT}*Supuestos!B11", NUM1,
     "Por encima del equilibrio", "Si cae por debajo de 'Firma'!B13, la firma pierde dinero ese año."),
    ("2", "Tasa de conversión de consulta", f"={CONV_CELL}", PCT,
     "40 % o más", "Revisar tiempo de respuesta, quién atiende y si se pide el cierre."),
    ("3", "Utilidad operativa", "=Firma!B9", MONEY,
     "Positiva y creciente", "Si baja con ingreso estable, el problema está en el costo de entrega."),
    ("4", "Margen de contribución (%)", "=Firma!B7", PCT,
     "45 % o más", "Si baja, mirar la columna 'Margen %' de TiposDeCaso: hay un producto subsidiado."),
    ("5", "Colchón sobre el punto de equilibrio", "=Firma!B16", PCT,
     "30 % o más", "Por debajo del 15 %, cualquier mes malo compromete la nómina."),
]
r = 5
for num, name, formula, fmt, ref, action in panel:
    ws.cell(row=r, column=1, value=num).font = Font(name=F, size=12, bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=name).font = BOLD
    c = ws.cell(row=r, column=3, value=formula)
    c.font = Font(name=F, size=12, bold=True, color=NAVY)
    c.number_format = fmt
    c.fill = PatternFill("solid", fgColor="E8EEF7")
    c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=4, value=ref).font = BLACK
    n = ws.cell(row=r, column=5, value=action)
    n.font = SUB
    n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 34
    r += 1

r += 1
ws.cell(row=r, column=2, value="Dos señales de alarma que se vigilan aparte").font = SECT
for t in [
    "Concentración de canal — si el 70 % de los casos viene de una sola fuente, esa fuente es un punto único de fallo.",
    "Días entre firma del caso y cobro íntegro — es el primer síntoma de casi cualquier problema de gestión, mucho antes de que se vea en la utilidad.",
]:
    r += 1
    c = ws.cell(row=r, column=2, value="•  " + t)
    c.font = SUB
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 24

out = "Modelo-Financiero-Santiago-PLLC.xlsx"
wb.save(out)
print("wrote", out)
